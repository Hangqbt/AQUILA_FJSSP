import pandas as pd
from scipy.stats import wilcoxon
import itertools
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

def generate_detailed_excel():
    # 1. Map algorithms to their specific CSV filenames
    files = {
        "AO Tabu (25/25/50)": "ao_benchmark_results_50prob.csv",
        "AO No Hybrid (25/25/50)": "ao_no_hybrid_benchmark_results_50prob.csv",
        "AO Tabu (60/20/20)": "ao_benchmark_results_20rand.csv",
        "AO No Hybrid (60/20/20)": "ao_benchmark_results_no_hybrid_20rand.csv"
    }

    raw_data = {}
    mean_data = {}

    # 2. Process files
    for name, filepath in files.items():
        df = pd.read_csv(filepath)
        df = df.sort_values('instance_name').set_index('instance_name')

        # Keep original mean ± std as strings
        raw_data[name] = df['objective_value']

        # Extract mean for mathematical differences
        mean_data[name] = df['objective_value'].apply(lambda x: float(str(x).split('±')[0].strip()))

    df_raw = pd.DataFrame(raw_data)
    df_means = pd.DataFrame(mean_data)

    # 3. Calculate Pairwise Differences PER INSTANCE
    diff_records = []
    for idx in df_means.index:
        row_diff = {"instance_name": idx}
        for col1, col2 in itertools.combinations(df_means.columns, 2):
            val1 = df_means.at[idx, col1]
            val2 = df_means.at[idx, col2]

            # Math: (B - A) / A
            pct_diff = ((val2 - val1) / val1)
            pair_name = f"{col2} vs {col1}"
            row_diff[pair_name] = pct_diff

        diff_records.append(row_diff)

    df_instance_diffs = pd.DataFrame(diff_records).set_index('instance_name')

    # 4. Calculate Overall Statistics (Wilcoxon)
    stats_results = []
    for col1, col2 in itertools.combinations(df_means.columns, 2):
        pct_diff_avg = ((df_means[col2] - df_means[col1]) / df_means[col1]).mean()
        stat, p_val = wilcoxon(df_means[col1], df_means[col2], zero_method='zsplit')
        stats_results.append({
            "Algorithm A": col1,
            "Algorithm B": col2,
            "Avg % Diff (A vs B)": float(pct_diff_avg),
            "Wilcoxon Stat": float(stat),
            "p-value": float(p_val),
            "Significant (p < 0.05)?": "Yes" if p_val < 0.05 else "No"
        })

    df_stats = pd.DataFrame(stats_results)

    # 5. Export to Excel with multiple sheets
    output_path = "Detailed_Init_Strategy_Analysis.xlsx"
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df_raw.reset_index().to_excel(writer, sheet_name='Raw Results (Mean ± Std)', index=False)
        df_instance_diffs.reset_index().to_excel(writer, sheet_name='Instance % Differences', index=False)
        df_stats.to_excel(writer, sheet_name='Statistical Summary', index=False)

    # 6. Apply formatting
    wb = openpyxl.load_workbook(output_path)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2F4F4F")
    border_thin = Border(left=Side(style='thin', color='D3D3D3'),
                         right=Side(style='thin', color='D3D3D3'),
                         top=Side(style='thin', color='D3D3D3'),
                         bottom=Side(style='thin', color='D3D3D3'))
    align_center = Alignment(horizontal="center", vertical="center")

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = align_center
            cell.border = border_thin
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.border = border_thin
                cell.alignment = align_center

    # Format percentages beautifully
    ws_diffs = wb['Instance % Differences']
    for row in ws_diffs.iter_rows(min_row=2, min_col=2):
        for cell in row:
            cell.number_format = '+0.00%;-0.00%;0.00%'

    ws_stats = wb['Statistical Summary']
    for row in ws_stats.iter_rows(min_row=2, min_col=3, max_col=3):
        for cell in row:
            cell.number_format = '+0.000%;-0.000%;0.000%'
    for row in ws_stats.iter_rows(min_row=2, min_col=5, max_col=5):
        for cell in row:
            cell.number_format = '0.0000'

    # Auto-adjust column widths
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for col in ws.columns:
            max_length = max((len(str(cell.value)) for cell in col), default=0)
            ws.column_dimensions[col[0].column_letter].width = min(max_length + 3, 40)

    # Freeze the instance names on the left
    wb['Raw Results (Mean ± Std)'].freeze_panes = 'B2'
    wb['Instance % Differences'].freeze_panes = 'B2'

    wb.save(output_path)
    print(f"Excel file successfully generated at: {output_path}")

if __name__ == "__main__":
    generate_detailed_excel()